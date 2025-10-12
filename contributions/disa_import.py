from . import models
from datetime import datetime
from django.conf import settings
from django.contrib import messages
from django.forms import ValidationError
from django.utils.translation import gettext_lazy as _
from licenses.models import License
from openpyxl import load_workbook
from zoneinfo import ZoneInfo
import logging
import re


logger = logging.getLogger('django')

BEGIN = 1
END = 2
DURATION = 3
TITLE = 4
TYPE = 11

WS_NAME = 'Auftragsfenster'
INFO = 'Infoblock'
LIVE = 'Live-Quelle'
IGNORED_PREFIXES = ['Trailer', 'Programmvorschau']


def _check_title(title: str, type: str) -> bool:
    """Check weather the title is valid."""
    if re.match(r'^\d+_', title):
        return True

    if type == INFO:
        return True

    if any([title.startswith(x) for x in IGNORED_PREFIXES]):
        return True

    return False


def validate(file):
    """
    Validate the DISA export file.

    Check weather the column and title naming is right.
    """
    wb = load_workbook(file)
    errors: list[ValidationError] = []

    def e(message: str) -> None:
        """Append an new ValidationError to the error list."""
        errors.append(ValidationError(message))

    if WS_NAME not in wb.sheetnames:
        e(_('The worksheet needs to be named "%(name)s".') % {'name': WS_NAME})
        raise ValidationError(errors)

    ws = wb[WS_NAME]
    rows = ws.rows
    header = next(rows)

    if header[BEGIN].value != (NAME := 'Anfang'):
        e(_('Column %(nr)s needs to be named %(name)s')
          % {
            'nr': BEGIN,
            'name': NAME
        })

    if header[END].value != (NAME := 'Ende'):
        e(_('Column %(nr)s needs to be named %(name)s')
          % {
            'nr': END,
            'name': NAME
        })

    if header[DURATION].value != (NAME := 'Länge'):
        e(_('Column %(nr)s needs to be named %(name)s')
          % {
            'nr': DURATION,
            'name': NAME
        })

    if header[TITLE].value != (NAME := 'Titel'):
        e(_('Column %(nr)s needs to be named %(name)s')
          % {
            'nr': TITLE,
            'name': NAME
        })

    if header[TYPE].value != (NAME := 'Typ'):
        e(_('Column %(nr)s needs to be named %(name)s')
          % {
            'nr': TYPE,
            'name': NAME
        })

    for row in rows:
        if not any([row[i].value for i in range(TYPE)]):
            continue
        if (not _check_title(row[TITLE].value, row[TYPE].value)):
            e(_('Invalid title in cell %(c)s%(r)s. Title needs the format'
                ' <nr>_<title>.') %
              {
                  'c': row[TITLE].column_letter,
                  'r': row[TITLE].row,
            })

    if errors:
        raise ValidationError(errors)


def disa_import(request, file):
    """
    Import contributions from DISA export - OPTIMIZED VERSION.

    All valid data gets imported even if an error occurs.
    """
    wb = load_workbook(file)
    ws = wb[WS_NAME]
    rows = ws.rows
    next(rows)  # ignore headers
    next(rows)  # ignore empty row

    # ОПТИМИЗАЦИЯ 1: Предзагрузка всех лицензий одним запросом
    all_license_numbers = set()
    rows_data = []
    
    for row in rows:
        if not any([row[i].value for i in range(TYPE)]):
            break

        if (
            row[TYPE].value == INFO or
            any([row[TITLE].value.startswith(x) for x in IGNORED_PREFIXES])
        ):
            continue

        nr = re.match(r'^\d+', row[TITLE].value)[0]
        all_license_numbers.add(nr)
        rows_data.append(row)
    
    # Один SQL запрос для всех лицензий вместо N запросов
    licenses_dict = {
        lic.number: lic 
        for lic in License.objects.filter(number__in=all_license_numbers).select_related()
    }
    
    # ОПТИМИЗАЦИЯ 2: Предзагрузка лицензий с запретом на повторы
    no_repetition_license_ids = set(
        License.objects.filter(
            number__in=all_license_numbers,
            repetitions_allowed=False
        ).values_list('id', flat=True)
    )
    
    # ОПТИМИЗАЦИЯ 3: Предзагрузка существующих contributions для лицензий с запретом
    existing_contributions = set(
        models.Contribution.objects.filter(
            license_id__in=no_repetition_license_ids
        ).values_list('license_id', flat=True).distinct()
    )

    dates = {}
    created_counter = 0
    
    for row in rows_data:
        nr = re.match(r'^\d+', row[TITLE].value)[0]

        license = licenses_dict.get(nr)
        if not license:
            msg = _('No license with number %(n)s found.') % {'n': nr}
            logger.error(msg)
            messages.error(request, msg)
            continue

        b_lst = list[int](re.split(r'\.| |:', row[BEGIN].value))
        b_lst = [int(x) for x in b_lst]

        broadcast_date = datetime(
            day=b_lst[0],
            month=b_lst[1],
            year=b_lst[2],
            hour=b_lst[3],
            minute=b_lst[4],
            second=b_lst[5],
            tzinfo=ZoneInfo(settings.TIME_ZONE)
        )

        # ОПТИМИЗАЦИЯ 4: Batch delete contributions по датам
        if not dates.get(hash(broadcast_date.date())):
            models.Contribution.objects.filter(
                broadcast_date__date=broadcast_date.date()).delete()
            dates[hash(broadcast_date.date())] = True

        # ОПТИМИЗАЦИЯ 5: Проверка без дополнительного SQL запроса
        if license.id in no_repetition_license_ids and license.id in existing_contributions:
            msg = _('No repetitions for number %(n)s allowed and already'
                    ' found a primary contribution.' % {'n': nr})
            logger.error(msg)
            messages.error(request, msg)
            continue

        contr, contr_created = models.Contribution.objects.get_or_create(
            license=license,
            broadcast_date=broadcast_date,
            live=(row[TYPE].value == LIVE),
        )

        if not contr_created:
            logger.warning(f'Contribution for license {nr} already exists.')
        else:
            created_counter += 1
            logger.info(f'Contribution {contr} created.')
            # Добавляем в кеш, чтобы следующая проверка работала
            if license.id in no_repetition_license_ids:
                existing_contributions.add(license.id)

    msg = _('Successfully created %d contributions.') % created_counter
    logger.info(msg)
    messages.info(request, msg)
