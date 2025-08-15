"""Module for importing inspection data from CSV and XLSX files."""

from .models import Inspection
from .models import InventoryItem
from datetime import datetime
from django.contrib import messages
from django.core.files import File
from django.core.files.temp import NamedTemporaryFile
from django.db import transaction
from django.forms import ValidationError
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl import load_workbook
from typing import Any
from typing import Dict
from typing import List
from typing import Optional
from typing import Set
from typing import Union
import csv
import logging


logger = logging.getLogger('django')
REQUIRED: Set[str] = {"inspection_number", "inspection_date"}


def _read_xlsx(file) -> List[Dict[str, str]]:
    """
    Read data from an XLSX file.

    Args:
        file: XLSX file object.

    Returns:
        List of dictionaries with row data.
    """
    wb = load_workbook(file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]  # type: ignore
    data = []
    for row in ws.iter_rows(min_row=2):  # type: ignore
        data.append({headers[i]: str(cell.value) if cell.value is not None else ''
                    for i, cell in enumerate(row)})
    return data


def validate(file: Any) -> None:
    """
    Validate the inspection file and check column names.

    Args:
        file: File object to validate.

    Raises:
        ValidationError: If the file format is invalid or required columns are missing.
    """
    file_extension = file.name.split('.')[-1].lower()
    headers: Optional[List[str]] = None

    if file_extension == 'csv':
        try:
            decoded_content = file.read().decode('utf-8')
        except UnicodeDecodeError:
            file.seek(0)
            decoded_content = file.read().decode('latin-1')
        file.seek(0)
        reader = csv.DictReader(decoded_content.splitlines())
        headers = reader.fieldnames
    elif file_extension == 'xlsx':
        wb = load_workbook(file)
        ws = wb.active
        if ws.max_row >= 1:  # type: ignore
            headers = [cell.value for cell in ws[1]]  # type: ignore
        else:
            headers = []
    else:
        raise ValidationError(_('File must be CSV or XLSX format'))

    if headers is None:
        raise ValidationError(_('Could not read headers from file.'))

    errors = []
    if not REQUIRED.issubset(set(headers)):
        missing_columns = REQUIRED - set(headers)
        errors.append(ValidationError(
            _('File must contain columns: %(columns)s. Missing: %(missing)s') % {
                'columns': ', '.join(REQUIRED),
                'missing': ', '.join(missing_columns)
            }
        ))

    if errors:
        raise ValidationError(errors)


def _parse_date(date_str: str) -> datetime.date:
    """
    Parse a date from various formats.

    Args:
        date_str: String with the date to parse.

    Returns:
        datetime.date object.

    Raises:
        ValueError: If the date format is not supported.
    """
    date_formats = [
        '%Y-%m-%d',  # 2024-03-20
        '%d.%m.%Y',  # 20.03.2024
        '%d/%m/%Y',  # 20/03/2024
        '%Y/%m/%d',  # 2024/03/20
    ]

    for date_format in date_formats:
        try:
            return datetime.strptime(date_str.strip(), date_format).date()
        except (ValueError, TypeError):
            continue

    raise ValueError(
        _('Invalid date format for "inspection_date". Supported formats: YYYY-MM-DD, DD.MM.YYYY, DD/MM/YYYY, YYYY/MM/DD. Received: "%(date_str)s"') % {'date_str': date_str}
    )


@transaction.atomic
def inspection_import(
    request: Optional[Any] = None,
    file: Optional[Any] = None,
    import_obj: Optional[Any] = None
) -> Dict[str, Union[int, str]]:
    """
    Import inspections from a CSV or XLSX file.

    Args:
        request: Django request object (can be None).
        file: File to import.
        import_obj: InspectionImport object for error logging.

    Returns:
        Dictionary with import results: {'created': int, 'skipped': int, 'error_log': str}.
    """
    created_counter = 0
    skipped_counter = 0
    error_logs: List[str] = []
    error_details: List[List[Any]] = []

    row_number_offset = 2  # Start from row 2 (row 1 is the header)

    try:
        if not file:
            raise ValueError(_("No file provided for import."))

        file.seek(0)
        file_extension = file.name.split('.')[-1].lower()

        rows_iterator: Union[csv.DictReader, List[Dict[str, str]]]
        header_row: Optional[List[str]] = None

        if file_extension == 'csv':
            try:
                decoded_content = file.read().decode('utf-8')
            except UnicodeDecodeError:
                file.seek(0)
                decoded_content = file.read().decode('latin-1')
            file.seek(0)
            csv_reader = csv.reader(decoded_content.splitlines())
            header_list = next(csv_reader, None)
            if header_list is None:
                raise ValueError(_("CSV file is empty or has no header row."))
            header_row = [h.strip() for h in header_list]
            rows_iterator = csv.DictReader(decoded_content.splitlines()[1:], fieldnames=header_row)
        elif file_extension == 'xlsx':
            rows_iterator = _read_xlsx(file)
            if rows_iterator:
                header_row = list(rows_iterator[0].keys()) if isinstance(rows_iterator[0], dict) else None
        else:
            raise ValidationError(_('Unsupported file format: %(ext)s') % {'ext': file_extension})

        if not header_row:
            raise ValueError(_("Could not determine headers from the file."))

        for i, row_data in enumerate(rows_iterator):
            current_row_num = i + row_number_offset
            try:
                for req_field in REQUIRED:
                    if req_field not in row_data or not row_data[req_field]:
                        raise ValueError(_("Missing required field: %(field)s") % {'field': req_field})

                inv_num = row_data.get("inventory_number", "").strip() or None
                item = InventoryItem.objects.filter(inventory_number=inv_num).first() if inv_num else None

                date_str = row_data.get("inspection_date", "")
                if not date_str:
                    raise ValueError(_("Field 'inspection_date' cannot be empty."))
                date_ = _parse_date(date_str)

                inspection_number = row_data.get("inspection_number", "").strip()
                if not inspection_number:
                    raise ValueError(_("Field 'inspection_number' cannot be empty."))

                defaults = {
                    "inventory_item": item,
                    "manufacturer": row_data.get("manufacturer", "").strip(),
                    "device_type": row_data.get("device_type", "").strip(),
                    "room": row_data.get("room", "").strip(),
                    "inspection_date": date_,
                    "result": row_data.get("result", "").strip(),
                    "target_part": row_data.get("target_part", "device").strip() or "device",
                }

                obj, is_created = Inspection.objects.get_or_create(
                    inspection_number=inspection_number,
                    defaults=defaults
                )

                if not is_created:
                    updated_fields = []
                    for field, value in defaults.items():
                        if getattr(obj, field) != value:
                            setattr(obj, field, value)
                            updated_fields.append(field)
                    if updated_fields:
                        obj.save(update_fields=updated_fields)
                    skipped_counter += 1
                else:
                    created_counter += 1

            except Exception as e:
                error_msg = str(e)
                error_logs.append(_('Row %(row_num)s: %(error)s') % {'row_num': current_row_num, 'error': error_msg})
                error_row_values = [row_data.get(h, '') for h in header_row]
                error_details.append([current_row_num, error_msg] + error_row_values)
                skipped_counter += 1
                if request:
                    messages.warning(request, _('Error processing row %(row_num)s: %(error)s') % {'row_num': current_row_num, 'error': error_msg})

        file.seek(0)

        if error_details and import_obj and header_row:
            error_wb = Workbook()
            ws = error_wb.active
            if ws:
                ws.title = _("Import Errors")
                ws.append([_('Row'), _('Error')] + header_row)
                for error_row_entry in error_details:
                    ws.append(error_row_entry)
                with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    error_wb.save(tmp.name)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = _('import_errors_%(timestamp)s.xlsx') % {
                        'timestamp': timestamp
                    }
                    if hasattr(import_obj, 'error_log_file'):
                        with open(tmp.name, 'rb') as tmp_file_to_save:
                            import_obj.error_log_file.save(filename, File(tmp_file_to_save))

    except Exception as e:
        logger.error(f"Inspection import failed: {str(e)}", exc_info=True)
        error_msg = str(e)
        if request:
            messages.error(request, _('Import failed: %(error)s') % {'error': error_msg})
        error_logs.append(_('Fatal Error: %(error)s') % {'error': error_msg})

    final_error_log = '\n'.join(error_logs) if error_logs else _('No errors during import.')
    if import_obj and hasattr(import_obj, 'error_log') and not getattr(import_obj, 'error_log_file', None):
        import_obj.error_log = final_error_log
        import_obj.save(update_fields=['error_log'])

    return {
        'created': created_counter,
        'skipped': skipped_counter,
        'error_log': final_error_log
    }
