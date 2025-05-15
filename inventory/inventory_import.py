from .models import InventoryItem
from .models import Manufacturer
from .models import Organization
from datetime import datetime
from django.contrib import messages
from django.core.files import File
from django.core.files.temp import NamedTemporaryFile
from django.forms import ValidationError
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
import re


logger = logging.getLogger('django')
WS_NAME = _('Inventory')
IGNORED_PREFIXES = [_('Test'), _('Sample')]


def _check_inventory_number(inventory_number: str) -> bool:
    """Check if the inventory number is valid."""
    if not re.match(r'^OK-\d+', inventory_number):
        return False
    return True


def _create_or_skip(model, field_name, value, row_number, request):
    """Create an object or skip the row if value is missing."""
    if value:
        obj, was_created = model.objects.get_or_create(**{field_name: value}, defaults={'description': ''})
        if was_created:
            logger.info(
                _('%(model)s "%(value)s" created automatically.') %
                {'model': model._meta.verbose_name, 'value': value}
            )
        return obj
    else:
        log_msg = _('%(model)s name is missing in row %(row)s. Skipping...') % {
            'model': model._meta.verbose_name, 'row': row_number
        }
        logger.warning(log_msg)
        messages.warning(request, log_msg)
        return None


def validate(file):
    """Validate the inventory file and check column naming."""
    wb = load_workbook(file)
    errors = []

    ws = wb.worksheets[0]
    rows = ws.rows
    header = next(rows)
    required_columns = {
        0: 'inventory_number',
        1: 'inventory_number_owner',
        2: 'description',
        3: 'serial_number',
        4: 'manufacturer',
        5: 'location',
        6: 'quantity',
        7: 'status',
        8: 'object_type',
        9: 'owner',
        10: 'purchase_date',
        11: 'purchase_cost',
        12: 'last_inspection'
    }

    # Check for all required columns
    for col_idx, expected_name in required_columns.items():
        actual_name = header[col_idx].value if col_idx < len(header) else None
        if actual_name != expected_name:
            errors.append(ValidationError(
                _('Column %(col)s should be named "%(expected)s", but got "%(actual)s"') % {
                    'col': col_idx + 1,
                    'expected': expected_name,
                    'actual': actual_name or 'missing'
                }
            ))

    if errors:
        raise ValidationError(errors)


def inventory_import(request, file, import_obj):
    """Import inventory items from an Excel file."""
    created_counter = 0
    skipped_counter = 0
    error_logs = []
    error_details = []

    try:
        wb = load_workbook(file)
        ws = wb.worksheets[0]
        rows = ws.rows
        headers = next(rows)

        for row in rows:
            try:
                inventory_number = str(row[0].value or '').strip()
                row_data = [str(cell.value) if cell.value is not None else '' for cell in row]

                if not inventory_number or not _check_inventory_number(inventory_number):
                    error_msg = _(f'Invalid inventory number "{inventory_number}"')
                    error_logs.append(f'Row {row[0].row}: {error_msg}')
                    error_details.append([row[0].row, error_msg] + row_data)
                    skipped_counter += 1
                    continue

                # Get values from row
                inventory_number_owner = str(row[1].value or '').strip() or None
                description = str(row[2].value or '').strip() or None
                serial_number = str(row[3].value or '').strip() or None
                manufacturer_name = str(row[4].value or '').strip() or None
                location = str(row[5].value or '').strip() or _('Unknown Location')

                # Check required fields
                if not location:
                    logger.warning(
                        _('Location is missing in row %(row)s. Skipping...') %
                        {'row': row[0].row}
                    )
                    messages.warning(
                        request,
                        _('Location is missing in row %(row)s. Skipping...') %
                        {'row': row[0].row}
                    )
                    skipped_counter += 1
                    continue

                # Handle quantity with default value 1
                try:
                    quantity = int(float(str(row[6].value or '1').strip()))
                except (ValueError, TypeError):
                    quantity = 1

                # Convert German status to English
                status_map = {
                    _('in Betrieb'): 'in_stock',
                    _('defekt'): 'defect',
                    _('ausgemustert'): 'written_off',
                    _('verliehen'): 'rented',
                    _('Ausleihe'): 'rented'
                }
                raw_status = str(row[7].value or '').strip()
                status = status_map.get(raw_status, 'in_stock')

                # Determine object type
                object_type = str(row[8].value or '').strip()
                if not object_type:
                    # If object type is not specified, use location value
                    object_type = location

                owner_name = str(row[9].value or '').strip() or None

                # Handle purchase date
                purchase_date = None
                if row[10].value:
                    try:
                        if isinstance(row[10].value, datetime):
                            purchase_date = row[10].value.date()
                        else:
                            purchase_date = datetime.strptime(str(row[10].value).strip(), '%Y-%m-%d').date()
                    except ValueError:
                        pass

                # Handle cost
                purchase_cost = None
                if row[11].value:
                    try:
                        purchase_cost = float(str(row[11].value).strip().replace(',', '.'))
                    except ValueError:
                        pass

                # Handle last_inspection
                last_inspection = str(row[12].value or '').strip() or None

                # Create or get manufacturer
                manufacturer = None
                if manufacturer_name:
                    manufacturer = _create_or_skip(Manufacturer, 'name', manufacturer_name, row[0].row, request)
                    if not manufacturer:
                        skipped_counter += 1
                        continue

                # Create or get owner
                owner = None
                if owner_name:
                    owner = _create_or_skip(Organization, 'name', owner_name, row[0].row, request)
                    if not owner:
                        skipped_counter += 1
                        continue

                # Check for duplicates
                if InventoryItem.objects.filter(inventory_number=inventory_number).exists():
                    logger.info(
                        _('Inventory item with number "%(number)s" already exists. Skipping...') %
                        {'number': inventory_number}
                    )
                    messages.warning(
                        request,
                        _('Inventory item with number "%(number)s" already exists. Skipping...') %
                        {'number': inventory_number}
                    )
                    skipped_counter += 1
                    continue

                # Create record
                InventoryItem.objects.create(
                    inventory_number=inventory_number,
                    inventory_number_owner=inventory_number_owner,
                    description=description,
                    serial_number=serial_number,
                    manufacturer=manufacturer,
                    location=location,
                    quantity=quantity,
                    status=status,
                    object_type=object_type,
                    owner=owner,
                    purchase_date=purchase_date,
                    purchase_cost=purchase_cost,
                    last_inspection=last_inspection
                )
                created_counter += 1

            except Exception as e:
                error_msg = str(e)
                error_logs.append(_('Row %(row)s: %(error)s') % {
                    'row': row[0].row,
                    'error': error_msg
                })
                error_details.append([row[0].row, error_msg] + row_data)
                skipped_counter += 1

        # If there are errors, create Excel file
        if error_details:
            error_wb = Workbook()
            ws = error_wb.active
            ws.title = _("Import Errors")

            # Headers
            ws.append([_('Row'), _('Error')] + [cell.value for cell in headers])

            # Error data
            for error_row in error_details:
                ws.append(error_row)

            # Save to temporary file
            with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                error_wb.save(tmp.name)
                # Save to model
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = _('import_errors_%(timestamp)s.xlsx') % {
                    'timestamp': timestamp
                }
                import_obj.error_log_file.save(filename, File(open(tmp.name, 'rb')))

    except Exception as e:
        error_msg = str(e)
        error_logs.append(_('Row %(row)s: %(error)s') % {
            'row': row[0].row,
            'error': error_msg
        })

    return {
        'created': created_counter,
        'skipped': skipped_counter,
        'error_log': '\n'.join(error_logs) if error_logs else _('No errors')
    }
